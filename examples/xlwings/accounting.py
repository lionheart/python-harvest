# accounting.py
import xlwings as xw
from xlwings import constants
from openpyxl import utils

from harvest import Harvest
from harvest.reports import Reports

reports = Reports("https://api.harvestapp.com/api/v2", put_auth_in_header=True, personal_token="Bearer ....", account_id=1234567890)

def get_harvest_invoices():
    wb = xw.Book.caller()

    sheets = xw.sheets

    invoice_sheet = None
    for sheet in sheets:
        if sheet.name == 'Harvest Invoices':
            invoice_sheet = sheet

    if invoice_sheet is not None:
        invoice_sheet.delete()

    new_invoices = sheets.add(name='Harvest Invoices', after='Sheet1')

    wb.app.screen_updating = False
    wb.app.calculation = 'manual'

    new_invoices.range('A1').value = 'client.id'
    new_invoices.range('B1').value = 'id'
    new_invoices.range('C1').value = 'number'
    new_invoices.range('D1').value = 'purchase_order'
    new_invoices.range('E1').value = 'tax'
    new_invoices.range('F1').value = 'tax_amount'
    new_invoices.range('G1').value = 'tax2'
    new_invoices.range('H1').value = 'tax2_amount'
    new_invoices.range('I1').value = 'discount'
    new_invoices.range('J1').value = 'discount_amount'
    new_invoices.range('K1').value = 'period_start'
    new_invoices.range('L1').value = 'period_end'
    new_invoices.range('M1').value = 'paid_date'
    new_invoices.range('N1').value = 'closed_at'
    new_invoices.range('O1').value = 'paid_at'
    new_invoices.range('P1').value = 'estimate'
    new_invoices.range('Q1').value = 'retainer'
    new_invoices.range('R1').value = 'sent_at'
    new_invoices.range('S1').value = 'notes'
    new_invoices.range('T1').value = 'client_key'
    new_invoices.range('U1').value = 'amount'
    new_invoices.range('V1').value = 'due_amount'
    new_invoices.range('W1').value = 'subject'
    new_invoices.range('X1').value = 'state'
    new_invoices.range('Y1').value = 'issue_date'
    new_invoices.range('Z1').value = 'due_date'
    new_invoices.range('AA1').value = 'payment_term'
    new_invoices.range('AB1').value = 'created_at'
    new_invoices.range('AC1').value = 'updated_at'
    new_invoices.range('AD1').value = 'currency'
    new_invoices.range('AE1').value = 'creator.id'
    new_invoices.range('AF1').value = 'creator.name'
    new_invoices.range('AG1').value = 'client.name'

    new_invoices.range('AH1').value = 'line_item.id'
    new_invoices.range('AI1').value = 'line_item.kind'
    new_invoices.range('AJ1').value = 'line_item.description'
    new_invoices.range('AK1').value = 'line_item.quantity'
    new_invoices.range('AL1').value = 'line_item.unit_price'
    new_invoices.range('AM1').value = 'line_item.amount'
    new_invoices.range('AN1').value = 'line_item.taxed'
    new_invoices.range('AO1').value = 'line_item.taxed2'

    new_invoices.range('AP1').value = 'line_item.project.id'
    new_invoices.range('AQ1').value = 'line_item.project.name'

    invoices = reports.invoices()

    row_counter = 2
    for invoice in invoices.invoices:
        for line_item in invoice.line_items:
            new_invoices.range('A%s' % row_counter).value = invoice.client.id
            new_invoices.range('B%s' % row_counter).value = invoice.id
            new_invoices.range('C%s' % row_counter).value = invoice.number
            new_invoices.range('D%s' % row_counter).value = invoice.purchase_order
            new_invoices.range('E%s' % row_counter).value = invoice.tax
            new_invoices.range('F%s' % row_counter).value = invoice.tax_amount
            new_invoices.range('G%s' % row_counter).value = invoice.tax2
            new_invoices.range('H%s' % row_counter).value = invoice.tax2_amount
            new_invoices.range('I%s' % row_counter).value = invoice.discount
            new_invoices.range('J%s' % row_counter).value = invoice.discount_amount
            new_invoices.range('K%s' % row_counter).value = invoice.period_start
            new_invoices.range('L%s' % row_counter).value = invoice.period_end
            new_invoices.range('M%s' % row_counter).value = invoice.paid_date
            new_invoices.range('N%s' % row_counter).value = invoice.closed_at
            new_invoices.range('O%s' % row_counter).value = invoice.paid_at
            new_invoices.range('P%s' % row_counter).value = invoice.estimate
            new_invoices.range('Q%s' % row_counter).value = invoice.retainer
            new_invoices.range('R%s' % row_counter).value = invoice.sent_at
            new_invoices.range('S%s' % row_counter).value = invoice.notes
            new_invoices.range('T%s' % row_counter).value = invoice.client_key
            new_invoices.range('U%s' % row_counter).value = invoice.amount
            new_invoices.range('V%s' % row_counter).value = invoice.due_amount
            new_invoices.range('W%s' % row_counter).value = invoice.subject
            new_invoices.range('X%s' % row_counter).value = invoice.state
            new_invoices.range('Y%s' % row_counter).value = invoice.issue_date
            new_invoices.range('Z%s' % row_counter).value = invoice.due_date
            new_invoices.range('AA%s' % row_counter).value = invoice.payment_term
            new_invoices.range('AB%s' % row_counter).value = invoice.created_at
            new_invoices.range('AC%s' % row_counter).value = invoice.updated_at
            new_invoices.range('AD%s' % row_counter).value = invoice.currency
            new_invoices.range('AE%s' % row_counter).value = invoice.creator.id
            new_invoices.range('AF%s' % row_counter).value = invoice.creator.name
            new_invoices.range('AG%s' % row_counter).value = invoice.client.name

            new_invoices.range('AH%s' % row_counter).value = line_item.id
            new_invoices.range('AI%s' % row_counter).value = line_item.kind
            new_invoices.range('AJ%s' % row_counter).value = line_item.description
            new_invoices.range('AK%s' % row_counter).value = line_item.quantity
            new_invoices.range('AL%s' % row_counter).value = line_item.unit_price
            new_invoices.range('AM%s' % row_counter).value = line_item.amount
            new_invoices.range('AN%s' % row_counter).value = line_item.taxed
            new_invoices.range('AO%s' % row_counter).value = line_item.taxed2

            if line_item.project is not None:
                new_invoices.range('AP%s' % row_counter).value = line_item.project.id
                new_invoices.range('AQ%s' % row_counter).value = line_item.project.name

            row_counter += 1

    wb.app.screen_updating = True
    wb.app.calculation = 'automatic'


def get_harvest_time_entries():
    wb = xw.Book.caller()

    sheets = xw.sheets

    time_entries_sheet = None
    for sheet in sheets:
        if sheet.name == 'Harvest Time Entries':
            time_entries_sheet = sheet

    if time_entries_sheet is not None:
        time_entries_sheet.delete()

    new_time_entries = sheets.add(name='Harvest Time Entries', after='Sheet1')

    wb.app.screen_updating = False
    wb.app.calculation = 'manual'

    new_time_entries.range('A1').value = 'notes' # OPTIONAL
    new_time_entries.range('B1').value = 'locked_reason' # OPTIONAL
    new_time_entries.range('C1').value = 'timer_started_at' # OPTIONAL
    new_time_entries.range('D1').value = 'started_time' # OPTIONAL
    new_time_entries.range('E1').value = 'ended_time' # OPTIONAL
    new_time_entries.range('F1').value = 'invoice' #COMPUND OBJECT # OPTIONAL
    new_time_entries.range('G1').value = 'external_reference' # OPTIONAL
    new_time_entries.range('H1').value = 'external_reference' # OPTIONAL
    new_time_entries.range('I1').value = 'billable_rate' # OPTIONAL
    new_time_entries.range('J1').value = 'id'
    new_time_entries.range('K1').value = 'spent_date'
    new_time_entries.range('L1').value = 'user.id'
    new_time_entries.range('M1').value = 'user.name'
    new_time_entries.range('N1').value = 'client.id'
    new_time_entries.range('O1').value = 'client.name'
    new_time_entries.range('P1').value = 'project.id'
    new_time_entries.range('Q1').value = 'project.name'
    new_time_entries.range('R1').value = 'task.id'
    new_time_entries.range('S1').value = 'task.name'
    new_time_entries.range('T1').value = 'user_assignment.id'
    new_time_entries.range('U1').value = 'user_assignment.is_project_manager'
    new_time_entries.range('V1').value = 'user_assignment.is_active'
    new_time_entries.range('W1').value = 'user_assignment.budget'
    new_time_entries.range('X1').value = 'user_assignment.created_at'
    new_time_entries.range('Y1').value = 'user_assignment.updated_at'
    new_time_entries.range('Z1').value = 'user_assignment.hourly_rate'
    new_time_entries.range('AA1').value = 'task_assignment.id'
    new_time_entries.range('AB1').value = 'task_assignment.billable'
    new_time_entries.range('AC1').value = 'task_assignment.is_active'
    new_time_entries.range('AD1').value = 'task_assignment.created_at'
    new_time_entries.range('AE1').value = 'task_assignment.updated_at'
    new_time_entries.range('AF1').value = 'task_assignment.jourly_rate'
    new_time_entries.range('AG1').value = 'task_assignment.budget'
    new_time_entries.range('AH1').value = 'hours'
    new_time_entries.range('AI1').value = 'created_at'
    new_time_entries.range('AJ1').value = 'is_locked'
    new_time_entries.range('AK1').value = 'is_closed'
    new_time_entries.range('AL1').value = 'is_billed'
    new_time_entries.range('AM1').value = 'is_running'
    new_time_entries.range('AN1').value = 'billable'
    new_time_entries.range('AO1').value = 'budgeted'
    new_time_entries.range('AP1').value = 'cost_rate'

    time_entries = reports.time_entries()

    row_counter = 2
    for time_entry in time_entries.time_entries:
        new_time_entries.range('A%s' % row_counter).value = time_entry.notes # OPTIONAL
        new_time_entries.range('B%s' % row_counter).value = time_entry.locked_reason # OPTIONAL
        new_time_entries.range('C%s' % row_counter).value = time_entry.timer_started_at # OPTIONAL
        new_time_entries.range('D%s' % row_counter).value = time_entry.started_time # OPTIONAL
        new_time_entries.range('E%s' % row_counter).value = time_entry.ended_time # OPTIONAL
        new_time_entries.range('F%s' % row_counter).value = time_entry.invoice #COMPUND OBJECT # OPTIONAL
        new_time_entries.range('G%s' % row_counter).value = time_entry.external_reference # OPTIONAL
        new_time_entries.range('H%s' % row_counter).value = time_entry.external_reference # OPTIONAL
        new_time_entries.range('I%s' % row_counter).value = time_entry.billable_rate # OPTIONAL
        new_time_entries.range('J%s' % row_counter).value = time_entry.id
        new_time_entries.range('K%s' % row_counter).value = time_entry.spent_date
        new_time_entries.range('L%s' % row_counter).value = time_entry.user.id
        new_time_entries.range('M%s' % row_counter).value = time_entry.user.name
        new_time_entries.range('N%s' % row_counter).value = time_entry.client.id
        new_time_entries.range('O%s' % row_counter).value = time_entry.client.name
        new_time_entries.range('P%s' % row_counter).value = time_entry.project.id
        new_time_entries.range('Q%s' % row_counter).value = time_entry.project.name
        new_time_entries.range('R%s' % row_counter).value = time_entry.task.id
        new_time_entries.range('S%s' % row_counter).value = time_entry.task.name
        new_time_entries.range('T%s' % row_counter).value = time_entry.user_assignment.id
        new_time_entries.range('U%s' % row_counter).value = time_entry.user_assignment.is_project_manager
        new_time_entries.range('V%s' % row_counter).value = time_entry.user_assignment.is_active
        new_time_entries.range('W%s' % row_counter).value = time_entry.user_assignment.budget
        new_time_entries.range('X%s' % row_counter).value = time_entry.user_assignment.created_at
        new_time_entries.range('Y%s' % row_counter).value = time_entry.user_assignment.updated_at
        new_time_entries.range('Z%s' % row_counter).value = time_entry.user_assignment.hourly_rate
        new_time_entries.range('AA%s' % row_counter).value = time_entry.task_assignment.id
        new_time_entries.range('AB%s' % row_counter).value = time_entry.task_assignment.billable
        new_time_entries.range('AC%s' % row_counter).value = time_entry.task_assignment.is_active
        new_time_entries.range('AD%s' % row_counter).value = time_entry.task_assignment.created_at
        new_time_entries.range('AE%s' % row_counter).value = time_entry.task_assignment.updated_at
        new_time_entries.range('AF%s' % row_counter).value = time_entry.task_assignment.hourly_rate
        new_time_entries.range('AG%s' % row_counter).value = time_entry.task_assignment.budget
        new_time_entries.range('AH%s' % row_counter).value = time_entry.hours
        new_time_entries.range('AI%s' % row_counter).value = time_entry.created_at
        new_time_entries.range('AJ%s' % row_counter).value = time_entry.is_locked
        new_time_entries.range('AK%s' % row_counter).value = time_entry.is_closed
        new_time_entries.range('AL%s' % row_counter).value = time_entry.is_billed
        new_time_entries.range('AM%s' % row_counter).value = time_entry.is_running
        new_time_entries.range('AN%s' % row_counter).value = time_entry.billable
        new_time_entries.range('AO%s' % row_counter).value = time_entry.budgeted
        new_time_entries.range('AP%s' % row_counter).value = time_entry.cost_rate

        row_counter += 1

    wb.app.screen_updating = True
    wb.app.calculation = 'automatic'

def create_harvest_invoices(client_id, invoice):
    new_invoice = reports.create_invoice(client_id, invoice)


def delete_columns(sheet_to_delete_from, columns_to_keep, sheet_rename, delete=True):
    if not isinstance(columns_to_keep, list):
        return

    # get the worksheet we are interested in (providing it exists)
    wb = xw.Book.caller()
    sheets = xw.sheets
    working_sheet = None
    for sheet in sheets:
        if sheet.name == sheet_to_delete_from:
            working_sheet = sheet

    if working_sheet is not None:
        columns = working_sheet.used_range.columns.count

        # excel column references are 1 based
        columns += 1
        length = 0
        column_ordinals_to_delete = []

        for col in range(1, columns):
            length += 1
            # are we getting rid of these columns, or keeping them and deleting the others..?
            if delete:
                if working_sheet.cells(1, col).value not in columns_to_keep:
                    column_ordinals_to_delete.append(length)
            else:
                if working_sheet.cells(1, col).value in columns_to_keep:
                    column_ordinals_to_delete.append(length)

        # delete from the right-hand-side first so the ordinals don't change.
        column_ordinals_to_delete.reverse()

        for column in column_ordinals_to_delete:
            letter = utils.cell.get_column_letter(column)
            working_sheet.range('%s:%s' % (letter, letter)).api.Delete(constants.DeleteShiftDirection.xlShiftToLeft)

        working_sheet.name = sheet_rename

def keep_columns(sheet_to_delete_from, columns_to_delete, sheet_rename):
    return delete_columns(sheet_to_delete_from, columns_to_delete, sheet_rename, delete=False)


def employee_rate_and_total(sheet_range = 'Employee Hours', sheet_lookup = 'Employee Rates'):

    # get the worksheet we are interested in (providing it exists)
    wb = xw.Book.caller()
    sheets = xw.sheets
    working_sheet = None
    for sheet in sheets:
        if sheet.name == sheet_range:
            working_sheet = sheet

    if working_sheet is not None:
        Lookup_cols = working_sheet.used_range.columns.count
        Lookup_rows = working_sheet.used_range.rows.count + 1

        rates_col = Lookup_cols + 1                   #calculates the column number for the rates
        total_col = Lookup_cols + 2                    #calculates the row number for the rates
        rows = working_sheet.used_range.rows.count             #calculates the number of rows in "Sheet1"
        code_position = 0
        surname_position = 0
        hours_position = 0

        for col in range(1, Lookup_cols + 1):
            if working_sheet.cells(1, col).value == 'Project Code':
                code_position = col

            if working_sheet.cells(1, col).value == 'Last Name':
                surname_position = col

            if working_sheet.cells(1, col).value == 'Hours':
                hours_position = col

        rates_Letter = utils.cell.get_column_letter(rates_col)
        total_Letter = utils.cell.get_column_letter(total_col)
        code_Letter = utils.cell.get_column_letter(code_position)
        surname_Letter = utils.cell.get_column_letter(surname_position)
        hours_Letter = utils.cell.get_column_letter(hours_position)

        working_sheet.cells(1,rates_col).value = "rates"
        working_sheet.cells(2,rates_col).value = "=VLOOKUP(%s%s, '%s'!$A$1:$B$%s, 2, FALSE)" % (surname_Letter, 2, sheet_lookup, Lookup_rows)
        working_sheet.cells(2,rates_col).api.autofill(working_sheet.range("%s2:%s%s" % (rates_Letter, rates_Letter, rows)).api, constants.AutoFillType.xlFillDefault)

        working_sheet.cells(1,total_col).value = "total"
        working_sheet.cells(2,total_col).value = "=%s2*%s2" %(hours_Letter, rates_Letter)
        working_sheet.cells(2,total_col).api.autofill(working_sheet.range("%s2:%s%s" % (total_Letter, total_Letter, rows)).api, constants.AutoFillType.xlFillDefault)

def make_employee_pivot_table():
    # get the worksheet we are interested in (providing it exists)
    wb = xw.Book.caller()
    sheets = xw.sheets
    working_sheet = None
    for sheet in sheets:
        if sheet.name == 'Employee Hours':
            working_sheet = sheet

    for col in range(1, working_sheet.used_range.columns.count + 1):
        if working_sheet.cells(1, col).value == 'Project Code':
            code_position = col

        if working_sheet.cells(1, col).value == 'Last Name':
            surname_position = col

        if working_sheet.cells(1, col).value == 'total':
            total_position = col

    code_Letter = utils.cell.get_column_letter(code_position)
    surname_Letter = utils.cell.get_column_letter(surname_position)
    total_Letter = utils.cell.get_column_letter(total_position)

    pivot_data = xw.sheets.add(name='Pivot Data', after='Sheet1')

    if working_sheet is not None:
        working_sheet.range("%s:%s" % (code_Letter, code_Letter)).api.Copy(pivot_data.range("A:A").api )
        working_sheet.range("%s:%s" % (surname_Letter, surname_Letter)).api.Copy(pivot_data.range("B:B").api )

        working_sheet.range("%s:%s" % (total_Letter, total_Letter)).api.Copy()
        pivot_data.range("C1").api.PasteSpecial(Paste = constants.PasteType.xlPasteValues)

        val_rows = pivot_data.used_range.rows.count                             #counts the number of rows in the values sheet

        PivotSourceRange = pivot_data.range("A1:C%s" % val_rows)                #selects the source data for the pivot table

        pivot_table = xw.sheets.add(name='Pivot Table', after='Sheet1')

        PivotTableName = 'ReportPivotTable'

        PivotCache = wb.api.PivotCaches().Create(SourceType=constants.PivotTableSourceType.xlDatabase, SourceData=PivotSourceRange.api, Version=constants.PivotTableVersionList.xlPivotTableVersion14)
        PivotTable = PivotCache.CreatePivotTable(TableDestination="'Pivot Table'!R1C1", TableName=PivotTableName, DefaultVersion=constants.PivotTableVersionList.xlPivotTableVersion14)
        PivotTable.PivotFields('Last Name').Orientation = constants.PivotFieldOrientation.xlRowField
        PivotTable.PivotFields('Last Name').Position = 1
        PivotTable.PivotFields('Project Code').Orientation = constants.PivotFieldOrientation.xlRowField
        PivotTable.PivotFields('Project Code').Position = 2
        PivotTable.PivotFields("total").Orientation = constants.PivotFieldOrientation.xlDataField

        piv_rows = pivot_table.used_range.rows.count                             #counts the number of rows in the pivot table

        people = []
        for each in PivotTable.PivotFields("Last Name").PivotItems():     #creates a list containing the names of all the people in the pivot table
            people.append(each.value)

        the_name = None
        for row in range(1, piv_rows):                                   #matches the name to the corresping code
            if (pivot_table.cells(row, 1).value in people):
                the_name = pivot_table.cells(row, 1).value
            else:
                pivot_table.cells(row, 3).value = the_name

        people = ["Row Labels", "Grand Total"] + people                                #adds 'Row Labels' and 'Grand Total' to the start and finish of the people list

        pivot_text = xw.sheets.add(name='Pivot Text', after='Sheet1')
        pivot_text.cells(1, 1).value = 'Project Code'
        pivot_text.cells(1, 2).value = 'total'
        pivot_text.cells(1, 3).value = 'Last Name'

        #copies and pastes the the pivot table to the pivot_text sheet
        pivot_table.range("A1:C%s" % piv_rows).api.Copy()
        pivot_text.range("A2").api.PasteSpecial(Paste = constants.PasteType.xlPasteValues)

        for row in reversed( range(1, piv_rows + 1) ):
            if pivot_text.cells(row, 1).value in people:
                pivot_text.cells(row, 1).api.EntireRow.Delete()

def project_code_and_names(sheet_range = 'Pivot Text', sheet_lookup = 'Project Code and Names'):

    # get the worksheet we are interested in (providing it exists)
    wb = xw.Book.caller()
    sheets = xw.sheets
    working_sheet = None
    for sheet in sheets:
        if sheet.name == sheet_range:
            working_sheet = sheet

    if working_sheet is not None:
        working_sheet.cells(1, 4).value = 'Project Code and Name'

        rows = working_sheet.used_range.rows.count             #calculates the number of rows in "Sheet1"

        working_sheet.cells(2,4).value = "=VLOOKUP(A2, '%s'!$A:$B, 2, FALSE)" % (sheet_lookup)
        working_sheet.cells(2,4).api.autofill(working_sheet.range("$D$2:$D$%s" % rows).api, constants.AutoFillType.xlFillDefault)


def time_report(**kwargs):
    wb = xw.Book.caller()

    sheets = xw.sheets

    time_entries_sheet = None
    for sheet in sheets:
        if sheet.name == 'Detailed Time Entries':
            time_entries_sheet = sheet

    if time_entries_sheet is not None:
        time_entries_sheet.delete()

    new_time_entries = sheets.add(name='Detailed Time Entries', after='Sheet1')

    wb.app.screen_updating = False
    wb.app.calculation = 'manual'

    new_time_entries.range('A1').value = 'Date'
    new_time_entries.range('B1').value = 'Client'
    new_time_entries.range('C1').value = 'Project'
    new_time_entries.range('D1').value = 'Project Code'
    new_time_entries.range('E1').value = 'Task'
    new_time_entries.range('F1').value = 'Notes'
    new_time_entries.range('G1').value = 'Hours'
    new_time_entries.range('H1').value = 'Billable?'
    new_time_entries.range('I1').value = 'Invoiced?'
    new_time_entries.range('J1').value = 'Approved?'
    new_time_entries.range('K1').value = 'First Name'
    new_time_entries.range('L1').value = 'Last Name'
    new_time_entries.range('M1').value = 'Roles'
    new_time_entries.range('N1').value = 'Employee?'
    new_time_entries.range('O1').value = 'Billable Rate'
    new_time_entries.range('P1').value = 'Billable Amount'
    new_time_entries.range('Q1').value = 'Cost Rate'
    new_time_entries.range('R1').value = 'Cost Amount'
    new_time_entries.range('S1').value = 'Currency'
    new_time_entries.range('T1').value = 'External Reference URL'

    detailed_time_entries = reports.detailed_time(**kwargs)

    row_counter = 2
    for detailed_time_entry in detailed_time_entries.detailed_time_entries:
        new_time_entries.range('A%s' % row_counter).value = detailed_time_entry.date
        new_time_entries.range('B%s' % row_counter).value = detailed_time_entry.client
        new_time_entries.range('C%s' % row_counter).value = detailed_time_entry.project
        new_time_entries.range('D%s' % row_counter).value = detailed_time_entry.project_code
        new_time_entries.range('E%s' % row_counter).value = detailed_time_entry.task
        new_time_entries.range('F%s' % row_counter).value = detailed_time_entry.notes
        new_time_entries.range('G%s' % row_counter).value = detailed_time_entry.hours
        new_time_entries.range('H%s' % row_counter).value = detailed_time_entry.billable
        new_time_entries.range('I%s' % row_counter).value = detailed_time_entry.invoiced
        new_time_entries.range('J%s' % row_counter).value = detailed_time_entry.approved
        new_time_entries.range('K%s' % row_counter).value = detailed_time_entry.first_name
        new_time_entries.range('L%s' % row_counter).value = detailed_time_entry.last_name
        new_time_entries.range('M%s' % row_counter).value = detailed_time_entry.roles
        new_time_entries.range('N%s' % row_counter).value = detailed_time_entry.employee
        new_time_entries.range('O%s' % row_counter).value = detailed_time_entry.billable_rate
        new_time_entries.range('P%s' % row_counter).value = detailed_time_entry.billable_amount
        new_time_entries.range('Q%s' % row_counter).value = detailed_time_entry.cost_rate
        new_time_entries.range('R%s' % row_counter).value = detailed_time_entry.cost_amount
        new_time_entries.range('S%s' % row_counter).value = detailed_time_entry.currency
        new_time_entries.range('T%s' % row_counter).value = detailed_time_entry.external_reference_url


        row_counter += 1

    wb.app.screen_updating = True
    wb.app.calculation = 'automatic'
